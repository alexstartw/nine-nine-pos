from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlmodel import Session

from ..database import get_session
from ..models import Product, StockEntry, StockEntryMethod, Vendor
from ..schemas import (
  PaginatedResponse,
  PaginationParams,
  ProductCreate,
  ProductImportRow,
  ProductImportSummary,
  ProductRead,
  ProductUpdate,
  ProductVendor,
)
from ..utils import calculate_gross_margin, generate_barcode

router = APIRouter(prefix='/products', tags=['products'])

IMPORT_HEADER_MAP = {
  '廠商': 'vendor_name',
  '廠商貨號': 'sku',
  '品名': 'name',
  '顏色': 'color',
  '尺寸': 'size',
  '進貨數量': 'quantity',
  '成本': 'cost',
  '售價': 'price'
}

REQUIRED_HEADERS = {'廠商', '廠商貨號', '品名', '成本', '進貨數量'}


def _product_to_read(product: Product, vendor: Vendor | None) -> ProductRead:
  gross, percent = calculate_gross_margin(product.price, product.cost)
  vendor_payload = ProductVendor.model_validate(vendor, from_attributes=True) if vendor else None
  return ProductRead.model_validate(
    product,
    from_attributes=True
  ).model_copy(update={'gross_margin': gross, 'gross_margin_percentage': percent, 'vendor': vendor_payload})


def _log_stock_entry(
  session: Session,
  product: Product,
  quantity: int,
  method: StockEntryMethod,
  vendor: Vendor | None = None,
  batch_id: str | None = None,
) -> None:
  if quantity <= 0:
    return
  if product.id is None:
    session.flush()

  entry = StockEntry(
    product_id=product.id,
    product_name=product.name,
    sku=product.sku,
    barcode=product.barcode,
    vendor_name=vendor.name if vendor else None,
    quantity=quantity,
    method=method,
    batch_id=batch_id,
  )
  session.add(entry)


@router.get('', response_model=PaginatedResponse[ProductRead])
def list_products(
  params: PaginationParams = Depends(),
  session: Session = Depends(get_session),
  q: Optional[str] = Query(default=None, description='依商品名稱、SKU 或條碼模糊搜尋'),
  vendor_id: Optional[int] = Query(default=None, description='指定廠商 ID'),
  first_stocked_from: Optional[datetime] = Query(default=None, description='第一次入庫開始時間'),
  first_stocked_to: Optional[datetime] = Query(default=None, description='第一次入庫結束時間'),
):
  filters = []
  if q:
    keyword = q.strip()
    if keyword:
      pattern = f'%{keyword}%'
      filters.append(or_(Product.name.ilike(pattern), Product.sku.ilike(pattern), Product.barcode.ilike(pattern)))
  if vendor_id:
    filters.append(Product.vendor_id == vendor_id)
  if first_stocked_from:
    filters.append(Product.first_stocked_at >= first_stocked_from)
  if first_stocked_to:
    filters.append(Product.first_stocked_at <= first_stocked_to)

  total_query = select(func.count()).select_from(Product)
  if filters:
    total_query = total_query.where(*filters)
  total = session.exec(total_query).scalar_one()

  statement = (
    select(Product, Vendor)
    .join(Vendor, Product.vendor_id == Vendor.id, isouter=True)
    .order_by(Product.updated_at.desc())
    .offset(params.offset)
    .limit(params.size)
  )
  if filters:
    statement = statement.where(*filters)

  rows = session.exec(statement).all()
  data = [_product_to_read(product, vendor) for product, vendor in rows]
  return PaginatedResponse[ProductRead](data=data, total=total, page=params.page, size=params.size)


@router.post('', response_model=ProductRead, status_code=status.HTTP_201_CREATED)
def create_product(
  payload: ProductCreate,
  session: Session = Depends(get_session)
):
  vendor = session.get(Vendor, payload.vendor_id) if payload.vendor_id else None
  barcode = generate_barcode(payload.vendor_id, payload.sku, payload.cost, payload.color, payload.size)
  now = datetime.utcnow()
  product = Product(
    **payload.model_dump(),
    barcode=barcode,
    first_stocked_at=now,
    data_updated_at=now,
    last_stocked_at=now,
  )
  session.add(product)
  session.flush()
  _log_stock_entry(session, product, product.stock, StockEntryMethod.SINGLE, vendor)
  session.commit()
  session.refresh(product)
  return _product_to_read(product, vendor)


@router.get('/{product_id}', response_model=ProductRead)
def get_product(product_id: int, session: Session = Depends(get_session)):
  product = session.get(Product, product_id)
  if not product:
    raise HTTPException(status_code=404, detail='Product not found')
  vendor = session.get(Vendor, product.vendor_id) if product.vendor_id else None
  return _product_to_read(product, vendor)


@router.put('/{product_id}', response_model=ProductRead)
def update_product(
  product_id: int,
  payload: ProductUpdate,
  session: Session = Depends(get_session)
):
  product = session.get(Product, product_id)
  if not product:
    raise HTTPException(status_code=404, detail='Product not found')

  update_data = payload.model_dump(exclude_unset=True)
  previous_stock = product.stock
  for key, value in update_data.items():
    setattr(product, key, value)

  if any(key in update_data for key in {'vendor_id', 'sku', 'cost', 'color', 'size'}):
    product.barcode = generate_barcode(product.vendor_id, product.sku, product.cost, product.color, product.size)
  now = datetime.utcnow()
  product.updated_at = now
  product.data_updated_at = now
  if 'stock' in update_data and update_data['stock'] is not None and product.stock > previous_stock:
    product.last_stocked_at = now

  session.add(product)
  session.commit()
  session.refresh(product)
  vendor = session.get(Vendor, product.vendor_id) if product.vendor_id else None
  return _product_to_read(product, vendor)


@router.delete('/{product_id}', status_code=status.HTTP_204_NO_CONTENT)
def delete_product(product_id: int, session: Session = Depends(get_session)):
  product = session.get(Product, product_id)
  if not product:
    raise HTTPException(status_code=404, detail='Product not found')
  session.delete(product)
  session.commit()
  return None


def _parse_import_rows(file_bytes: bytes) -> List[ProductImportRow]:
  try:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
  except Exception as exc:  # pragma: no cover - openpyxl specific error
    raise HTTPException(status_code=400, detail=f'無法讀取 Excel 檔案: {exc}')

  sheet = workbook.active
  headers: List[str] = []
  for cell in sheet[1]:
    headers.append(str(cell.value).strip() if cell.value is not None else '')

  missing = [header for header in REQUIRED_HEADERS if header not in headers]
  if missing:
    raise HTTPException(status_code=400, detail=f'欄位缺少: {", ".join(missing)}')

  header_index: Dict[str, int] = {name: headers.index(name) for name in IMPORT_HEADER_MAP if name in headers}

  def require_str(value, field: str, row_idx: int) -> str:
    if value is None:
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 必須為文字')
    text = str(value).strip()
    if not text:
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 不可為空')
    return text

  def require_int(value, field: str, row_idx: int, positive: bool = False) -> int:
    if value is None or str(value).strip() == '':
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 必須為整數')
    try:
      number = int(float(value))
    except (TypeError, ValueError) as exc:
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 必須為整數') from exc
    if positive and number <= 0:
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 必須大於 0')
    return number

  rows: List[ProductImportRow] = []
  for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    if all(cell is None or str(cell).strip() == '' for cell in row):
      continue

    data: Dict[str, str] = {}
    for header, key in IMPORT_HEADER_MAP.items():
      if header not in header_index:
        continue
      col_idx = header_index[header]
      value = row[col_idx] if col_idx < len(row) else None
      data[key] = value

    try:
      payload = ProductImportRow(
        vendor_name=require_str(data.get('vendor_name'), '廠商', idx),
        sku=require_str(data.get('sku'), '廠商貨號', idx),
        name=require_str(data.get('name'), '品名', idx),
        color=require_str(data.get('color'), '顏色', idx),
        size=require_str(data.get('size'), '尺寸', idx),
        cost=require_int(data.get('cost'), '成本', idx, positive=True),
        price=require_int(data.get('price'), '售價', idx, positive=True),
        quantity=require_int(data.get('quantity'), '進貨數量', idx, positive=True)
      )
    except Exception as exc:
      raise HTTPException(status_code=400, detail=f'列 {idx} 解析失敗: {exc}')

    rows.append(payload)

  if not rows:
    raise HTTPException(status_code=400, detail='檔案內沒有可處理的資料')

  return rows


@router.post('/import', response_model=ProductImportSummary, status_code=status.HTTP_201_CREATED)
async def import_products(
  file: UploadFile = File(...),
  session: Session = Depends(get_session)
):
  if not file.filename.lower().endswith(('.xlsx', '.xlsm')):
    raise HTTPException(status_code=400, detail='僅支援 .xlsx 或 .xlsm 檔案')

  file_bytes = await file.read()
  rows = _parse_import_rows(file_bytes)

  summary = ProductImportSummary()
  batch_id = f'import-{datetime.utcnow().strftime("%Y%m%d%H%M%S%f")}'

  for row in rows:
    if not row.vendor_name:
      summary.errors.append('缺少廠商名稱')
      continue

    vendor_row = session.exec(select(Vendor).where(Vendor.name == row.vendor_name)).first()
    vendor = None
    if vendor_row is not None:
      vendor = vendor_row[0] if not isinstance(vendor_row, Vendor) else vendor_row

    if not vendor:
      summary.errors.append(f"找不到廠商: {row.vendor_name}")
      continue

    if not row.sku:
      summary.errors.append('缺少廠商貨號')
      continue

    barcode = generate_barcode(vendor.id, row.sku, row.cost, row.color, row.size)
    product_row = session.exec(select(Product).where(Product.barcode == barcode)).first()
    product = None
    if product_row is not None:
      product = product_row[0] if not isinstance(product_row, Product) else product_row

    if product:
      now = datetime.utcnow()
      if product.first_stocked_at is None:
        product.first_stocked_at = now
      product.stock += row.quantity
      product.cost = row.cost
      product.price = row.price
      product.updated_at = now
      product.data_updated_at = now
      product.last_stocked_at = now
      session.add(product)
      summary.restocked += 1
      _log_stock_entry(session, product, row.quantity, StockEntryMethod.IMPORT, vendor, batch_id=batch_id)
    else:
      now = datetime.utcnow()
      product = Product(
        name=row.name or row.sku,
        sku=row.sku,
        vendor_id=vendor.id,
        barcode=barcode,
        color=row.color,
        size=row.size,
        cost=row.cost,
        price=row.price,
        stock=row.quantity,
        description=None,
        image_url=None,
        first_stocked_at=now,
        data_updated_at=now,
        last_stocked_at=now,
      )
      session.add(product)
      session.flush()
      _log_stock_entry(session, product, row.quantity, StockEntryMethod.IMPORT, vendor, batch_id=batch_id)
      summary.created += 1

  session.commit()

  return summary
