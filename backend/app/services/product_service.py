from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlmodel import Session

from ..models import OrderItem, Product, ReservationItem, StockEntry, StockEntryMethod, Vendor
from ..schemas import (
  LegacyProductImportRow,
  PaginatedResponse,
  ProductCreate,
  ProductImportRow,
  ProductImportSummary,
  ProductRead,
  ProductSummary,
  ProductUpdate,
  ProductVendor,
)
from ..utils import calculate_gross_margin, generate_barcode
from ..utils.time_utils import utc8_now

IMPORT_HEADER_MAP = {
  '廠商': 'vendor_name',
  '廠商貨號': 'sku',
  '品名': 'name',
  '顏色': 'color',
  '尺寸': 'size',
  '進貨數量': 'quantity',
  '成本': 'cost',
  '售價': 'price'
}
REQUIRED_HEADERS = {'廠商', '廠商貨號', '品名', '成本', '進貨數量'}

LEGACY_IMPORT_HEADER_MAP = {**IMPORT_HEADER_MAP, '條碼': 'barcode'}
LEGACY_REQUIRED_HEADERS = REQUIRED_HEADERS | {'條碼'}


class ProductService:
  def __init__(self, session: Session) -> None:
    self.session = session

  # ── Serialization ──────────────────────────────────────────────────────────

  def to_read(self, product: Product, vendor: Optional[Vendor]) -> ProductRead:
    gross, percent = calculate_gross_margin(product.price, product.cost)
    vendor_payload = ProductVendor.model_validate(vendor, from_attributes=True) if vendor else None
    return ProductRead.model_validate(
      product, from_attributes=True
    ).model_copy(update={'gross_margin': gross, 'gross_margin_percentage': percent, 'vendor': vendor_payload})

  # ── Query operations ───────────────────────────────────────────────────────

  def list(
    self,
    page: int,
    size: int,
    offset: int,
    q: Optional[str] = None,
    vendor_id: Optional[int] = None,
    first_stocked_from: Optional[datetime] = None,
    first_stocked_to: Optional[datetime] = None,
  ) -> PaginatedResponse[ProductRead]:
    filters = []
    if q:
      keyword = q.strip()
      if keyword:
        pattern = f'%{keyword}%'
        filters.append(or_(Product.name.ilike(pattern), Product.sku.ilike(pattern), Product.barcode.ilike(pattern)))
    if vendor_id:
      filters.append(Product.vendor_id == vendor_id)
    if first_stocked_from:
      filters.append(Product.first_stocked_at >= first_stocked_from)
    if first_stocked_to:
      filters.append(Product.first_stocked_at <= first_stocked_to)

    total_query = select(func.count()).select_from(Product)
    if filters:
      total_query = total_query.where(*filters)
    total = self.session.exec(total_query).scalar_one()

    statement = (
      select(Product, Vendor)
      .join(Vendor, Product.vendor_id == Vendor.id, isouter=True)
      .order_by(func.coalesce(Product.last_stocked_at, Product.updated_at).desc())
      .offset(offset)
      .limit(size)
    )
    if filters:
      statement = statement.where(*filters)

    rows = self.session.exec(statement).all()
    data = [self.to_read(product, vendor) for product, vendor in rows]
    return PaginatedResponse[ProductRead](data=data, total=total, page=page, size=size)

  def get_summary(self) -> ProductSummary:
    total_stock = self.session.exec(select(func.coalesce(func.sum(Product.stock), 0))).scalar_one()
    total_value = self.session.exec(
      select(func.coalesce(func.sum(Product.stock * Product.cost), 0))
    ).scalar_one()
    return ProductSummary(total_stock=int(total_stock or 0), total_stock_value=float(total_value or 0))

  def get_by_id(self, product_id: int) -> tuple[Product, Optional[Vendor]]:
    product = self.session.get(Product, product_id)
    if not product:
      raise HTTPException(status_code=404, detail='Product not found')
    vendor = self.session.get(Vendor, product.vendor_id) if product.vendor_id else None
    return product, vendor

  # ── Mutation operations ────────────────────────────────────────────────────

  def create(self, payload: ProductCreate) -> ProductRead:
    vendor = self.session.get(Vendor, payload.vendor_id) if payload.vendor_id else None
    now = utc8_now()
    product = Product(
      **payload.model_dump(),
      barcode=generate_barcode(payload.vendor_id, payload.sku, payload.cost, payload.color, payload.size),
      first_stocked_at=now,
      data_updated_at=now,
      last_stocked_at=now,
    )
    self.session.add(product)
    self.session.flush()
    self._log_stock_entry(product, product.stock, StockEntryMethod.SINGLE, vendor)
    self.session.commit()
    self.session.refresh(product)
    return self.to_read(product, vendor)

  def update(self, product_id: int, payload: ProductUpdate) -> ProductRead:
    product, _ = self.get_by_id(product_id)
    update_data = payload.model_dump(exclude_unset=True)
    previous_stock = product.stock
    for key, value in update_data.items():
      setattr(product, key, value)

    if any(key in update_data for key in {'vendor_id', 'sku', 'cost', 'color', 'size'}):
      product.barcode = generate_barcode(product.vendor_id, product.sku, product.cost, product.color, product.size)

    now = utc8_now()
    product.updated_at = now
    product.data_updated_at = now
    if 'stock' in update_data and update_data['stock'] is not None and product.stock > previous_stock:
      product.last_stocked_at = now

    self.session.add(product)
    self.session.commit()
    self.session.refresh(product)
    vendor = self.session.get(Vendor, product.vendor_id) if product.vendor_id else None
    return self.to_read(product, vendor)

  def delete(self, product_id: int) -> None:
    product, _ = self.get_by_id(product_id)

    has_orders = self.session.exec(
      select(func.count()).select_from(OrderItem).where(OrderItem.product_id == product_id)
    ).scalar_one()
    if has_orders:
      raise HTTPException(status_code=409, detail='此商品已有訂單銷售紀錄，無法刪除')

    has_active_reservations = self.session.exec(
      select(func.count()).select_from(ReservationItem).where(ReservationItem.product_id == product_id)
    ).scalar_one()
    if has_active_reservations:
      raise HTTPException(status_code=409, detail='此商品有預訂 / 保留紀錄，請先處理後再刪除')

    # nullify nullable FK in stock_entries before deleting product
    stock_entries = self.session.exec(
      select(StockEntry).where(StockEntry.product_id == product_id)
    ).scalars().all()
    for entry in stock_entries:
      entry.product_id = None
      self.session.add(entry)

    self.session.delete(product)
    self.session.commit()

  # ── Import operations ──────────────────────────────────────────────────────

  def import_from_excel(self, file_bytes: bytes) -> ProductImportSummary:
    rows = self._parse_import_rows(file_bytes)
    summary = ProductImportSummary()
    batch_id = f'import-{utc8_now(tz_aware=True).strftime("%Y%m%d%H%M%S%f")}'

    for row in rows:
      vendor = self._find_vendor(row.vendor_name)
      if not vendor:
        summary.errors.append(f'找不到廠商: {row.vendor_name}')
        continue

      barcode = generate_barcode(vendor.id, row.sku, row.cost, row.color, row.size)
      product = self._find_product_by_barcode(barcode)
      now = utc8_now()

      if product:
        if product.first_stocked_at is None:
          product.first_stocked_at = now
        product.stock += row.quantity
        product.cost = row.cost
        product.price = row.price
        product.updated_at = now
        product.data_updated_at = now
        product.last_stocked_at = now
        self.session.add(product)
        self._log_stock_entry(product, row.quantity, StockEntryMethod.IMPORT, vendor, batch_id=batch_id)
        summary.restocked += 1
      else:
        product = Product(
          name=row.name or row.sku, sku=row.sku, vendor_id=vendor.id,
          barcode=barcode, color=row.color, size=row.size, cost=row.cost,
          price=row.price, stock=row.quantity, first_stocked_at=now,
          data_updated_at=now, last_stocked_at=now,
        )
        self.session.add(product)
        self.session.flush()
        self._log_stock_entry(product, row.quantity, StockEntryMethod.IMPORT, vendor, batch_id=batch_id)
        summary.created += 1

    if summary.errors:
      self.session.rollback()
      raise HTTPException(
        status_code=422,
        detail={'errors': summary.errors, 'created': summary.created, 'restocked': summary.restocked},
      )

    self.session.commit()
    return summary

  def import_legacy_from_excel(self, file_bytes: bytes) -> ProductImportSummary:
    rows = self._parse_legacy_import_rows(file_bytes)
    summary = ProductImportSummary()
    batch_id = f'legacy-{utc8_now(tz_aware=True).strftime("%Y%m%d%H%M%S%f")}'

    for row in rows:
      vendor = self._find_vendor(row.vendor_name)
      if not vendor:
        summary.errors.append(f'找不到廠商: {row.vendor_name}')
        continue

      product = self._find_product_by_barcode(row.barcode)
      now = utc8_now()

      if product:
        if product.first_stocked_at is None:
          product.first_stocked_at = now
        product.vendor_id = vendor.id
        product.name = row.name or product.name
        product.sku = row.sku
        product.color = row.color
        product.size = row.size
        product.cost = row.cost
        product.price = row.price
        product.stock += row.quantity
        product.updated_at = now
        product.data_updated_at = now
        product.last_stocked_at = now
        self.session.add(product)
        self._log_stock_entry(product, row.quantity, StockEntryMethod.IMPORT, vendor, batch_id=batch_id)
        summary.restocked += 1
      else:
        product = Product(
          name=row.name or row.sku, sku=row.sku, vendor_id=vendor.id,
          barcode=row.barcode, color=row.color, size=row.size, cost=row.cost,
          price=row.price, stock=row.quantity, first_stocked_at=now,
          data_updated_at=now, last_stocked_at=now,
        )
        self.session.add(product)
        self.session.flush()
        self._log_stock_entry(product, row.quantity, StockEntryMethod.IMPORT, vendor, batch_id=batch_id)
        summary.created += 1

    if summary.errors:
      self.session.rollback()
      raise HTTPException(
        status_code=422,
        detail={'errors': summary.errors, 'created': summary.created, 'restocked': summary.restocked},
      )

    self.session.commit()
    return summary

  # ── Private helpers ────────────────────────────────────────────────────────

  def _log_stock_entry(
    self,
    product: Product,
    quantity: int,
    method: StockEntryMethod,
    vendor: Optional[Vendor] = None,
    batch_id: Optional[str] = None,
  ) -> None:
    if quantity <= 0:
      return
    if product.id is None:
      self.session.flush()
    now = utc8_now()
    self.session.add(StockEntry(
      product_id=product.id,
      product_name=product.name,
      sku=product.sku,
      barcode=product.barcode,
      vendor_name=vendor.name if vendor else None,
      quantity=quantity,
      method=method,
      batch_id=batch_id,
      created_at=now,
      updated_at=now,
    ))
    product.last_stocked_at = now
    self.session.add(product)

  def _find_vendor(self, name: str) -> Optional[Vendor]:
    row = self.session.exec(select(Vendor).where(Vendor.name == name)).first()
    if row is None:
      return None
    return row[0] if not isinstance(row, Vendor) else row

  def _find_product_by_barcode(self, barcode: str) -> Optional[Product]:
    row = self.session.exec(select(Product).where(Product.barcode == barcode)).first()
    if row is None:
      return None
    return row[0] if not isinstance(row, Product) else row

  @staticmethod
  def _require_str(value, field: str, row_idx: int) -> str:
    if value is None:
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 必須為文字')
    text = str(value).strip()
    if not text:
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 不可為空')
    return text

  @staticmethod
  def _require_int(value, field: str, row_idx: int, positive: bool = False) -> int:
    if value is None or str(value).strip() == '':
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 必須為整數')
    try:
      number = int(float(value))
    except (TypeError, ValueError) as exc:
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 必須為整數') from exc
    if positive and number <= 0:
      raise HTTPException(status_code=400, detail=f'列 {row_idx} 「{field}」 必須大於 0')
    return number

  def _parse_import_rows(self, file_bytes: bytes) -> List[ProductImportRow]:
    return self._parse_excel(file_bytes, IMPORT_HEADER_MAP, REQUIRED_HEADERS, is_legacy=False)

  def _parse_legacy_import_rows(self, file_bytes: bytes) -> List[LegacyProductImportRow]:
    return self._parse_excel(file_bytes, LEGACY_IMPORT_HEADER_MAP, LEGACY_REQUIRED_HEADERS, is_legacy=True)

  def _parse_excel(
    self,
    file_bytes: bytes,
    header_map: Dict[str, str],
    required_headers: set,
    is_legacy: bool
  ) -> list:
    try:
      workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # pragma: no cover
      raise HTTPException(status_code=400, detail=f'無法讀取 Excel 檔案: {exc}')

    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]

    missing = [h for h in required_headers if h not in headers]
    if missing:
      raise HTTPException(status_code=400, detail=f'欄位缺少: {", ".join(missing)}')

    header_index = {name: headers.index(name) for name in header_map if name in headers}
    rows = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
      if all(cell is None or str(cell).strip() == '' for cell in row):
        continue
      data = {
        key: (row[header_index[h]] if h in header_index and header_index[h] < len(row) else None)
        for h, key in header_map.items()
      }
      try:
        if is_legacy:
          rows.append(LegacyProductImportRow(
            vendor_name=self._require_str(data.get('vendor_name'), '廠商', idx),
            sku=self._require_str(data.get('sku'), '廠商貨號', idx),
            name=self._require_str(data.get('name'), '品名', idx),
            color=self._require_str(data.get('color'), '顏色', idx),
            size=self._require_str(data.get('size'), '尺寸', idx),
            cost=self._require_int(data.get('cost'), '成本', idx, positive=True),
            price=self._require_int(data.get('price'), '售價', idx, positive=True),
            quantity=self._require_int(data.get('quantity'), '進貨數量', idx, positive=True),
            barcode=self._require_str(data.get('barcode'), '條碼', idx),
          ))
        else:
          rows.append(ProductImportRow(
            vendor_name=self._require_str(data.get('vendor_name'), '廠商', idx),
            sku=self._require_str(data.get('sku'), '廠商貨號', idx),
            name=self._require_str(data.get('name'), '品名', idx),
            color=self._require_str(data.get('color'), '顏色', idx),
            size=self._require_str(data.get('size'), '尺寸', idx),
            cost=self._require_int(data.get('cost'), '成本', idx, positive=True),
            price=self._require_int(data.get('price'), '售價', idx, positive=True),
            quantity=self._require_int(data.get('quantity'), '進貨數量', idx, positive=True),
          ))
      except Exception as exc:
        raise HTTPException(status_code=400, detail=f'列 {idx} 解析失敗: {exc}')

    if not rows:
      raise HTTPException(status_code=400, detail='檔案內沒有可處理的資料')
    return rows
